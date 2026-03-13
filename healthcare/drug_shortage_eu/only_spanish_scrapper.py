import asyncio
import re, os
from playwright.async_api import async_playwright
from datetime import datetime
from openpyxl import Workbook, load_workbook


async def write_dict_to_xlsm(data_dict):
    output_dir = r"C:\Users\HP\Desktop"
    date_str = datetime.today().strftime("%Y%m%d")
    file_path = os.path.join(output_dir, f"Spain Output LL {date_str}.xlsx")

    # 👇 Define the fixed order of columns here
    column_order = [
        "Country",
        "ATC3",
        "ATC4",
        "Product Name",
        "Product Strength",
        "Product pack size",
        "Corporation",
        "Molecule",
        "Shortage Status",
        "Date reported",
        "Start Date",
        "End Date",
        "Date last updated",
        "Relevant unit numbers",
        "Alternatives available",
        "Shortage impact rating",
        "Additional Commentary"
    ]

    # Reorder data_dict according to column_order
    row = [data_dict.get(col, "") for col in column_order]

    if os.path.exists(file_path):
        wb = load_workbook(file_path)
        ws = wb.active
        ws.append(row)
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(column_order)  # Write header
        ws.append(row)

    wb.save(file_path)



async def auto_scroll(page, delay=300):
    previous_count = -1

    while True:
        # Get all visible items
        items = await page.query_selector_all("div#resultlist div.list-group-item.row div.titleDesabast")

        # If no new items appeared, stop
        if len(items) == previous_count:
            break

        previous_count = len(items)

        # Scroll each item into view to trigger lazy loading
        for item in items:
            try:
                await item.scroll_into_view_if_needed()
                await page.wait_for_timeout(delay)
            except:
                continue




async def extract_dates(item):
    start_date = ""
    end_date = ""

    # Get all listsValues divs
    values = await item.query_selector_all("div.listsValues")

    if len(values) >= 2:
        start_date = (await values[1].inner_text()).strip()
    if len(values) >= 3:
        end_date = (await values[2].inner_text()).strip()

    return start_date, end_date


async def extract_strength_from_title(title_text):
    # Regex pattern to match strengths like 10 MG/G, 100 mg/ml, 2 mg/0,625 mg
    pattern = re.compile(r"(\d+(?:[.,]\d+)?(?:\s*/\s*\d+(?:[.,]\d+)?)?\s*(?:mg|mg/g|mg/ml|g|mcg|µg|%))", re.I)
    matches = pattern.findall(title_text)
    return "; ".join([m.strip() for m in matches]) if matches else ""


async def extract_molecule(page):
    pactivos_tag = await page.query_selector("div#pactivosList")
    if not pactivos_tag:
        return ""

    # Get all <li> inside pactivosList
    li_tags = await pactivos_tag.query_selector_all("li")

    molecules = []
    for li in li_tags:
        text = (await li.inner_text()).strip()
        if text:
            molecules.append(text)

    return "; ".join(molecules)


async def extract_dosis(page):
    dosis_tag = await page.query_selector("div#dosis")
    if not dosis_tag:
        return ""

    li_tags = await dosis_tag.query_selector_all("li")

    dosis_list = []
    for li in li_tags:
        text = (await li.inner_text()).strip()
        if text:
            dosis_list.append(text)

    return "; ".join(dosis_list)


async def compute_date_reported(start_date_str):
    # Convert string to datetime
    try:
        start_date = datetime.strptime(start_date_str, "%d/%m/%Y")
    except:
        # Invalid or empty start date, fallback to today
        return datetime.today().strftime("%d/%m/%Y")

    today = datetime.today()

    if start_date <= today:
        return start_date.strftime("%d/%m/%Y")
    else:
        return today.strftime("%d/%m/%Y")


async def extract_atc(page):
    atc_tag = await page.query_selector("div#atcList")
    if not atc_tag:
        return {"ATC4": "", "ATC3": ""}

    li_tags = await atc_tag.query_selector_all("li")
    codes = []

    for li in li_tags:
        text = (await li.inner_text()).strip()
        if text:
            # Extract code part before space or dash
            code_match = re.match(r"([A-Z0-9]+)", text)
            if code_match:
                codes.append(code_match.group(1))

    atc4 = "; ".join(codes)
    atc4 = atc4.strip().split(";")[-1].strip()
    atc3 = atc4[:4] if atc4 else ""

    return atc4, atc3


async def get_text_after_last_comma(text):
    parts = text.rsplit(",", 1)
    if len(parts) == 2:
        return parts[1].strip()
    else:
        return ""  # No comma found


async def scrape_detail(page, master_dict):
    # Product Name
    name_tag = await page.query_selector("h1#nombreMedicamento")
    product_name = (await name_tag.inner_text()).strip() if name_tag else ""

    # Product Strength
    title_text = product_name.strip() if product_name else ""
    strength = await extract_strength_from_title(title_text)

    # Company name
    company_tag = await page.query_selector("div#nombrelab")
    company_name = (await company_tag.inner_text()).strip() if company_tag else ""

    # Molecule
    molecule = await extract_molecule(page)

    # ATC
    atc4, atc3 = await extract_atc(page)

    #Shortage Status
    shortage_status = "Resolved"
    if master_dict["end_date"]:
        shortage_status = "Active"

    # Report date
    report_date = await compute_date_reported(master_dict["start_date"])

    data = {
        "Country": "Spain",
        "ATC3": atc3,
        "ATC4": atc4,
        "Product Name": product_name,
        "Product Strength": strength,
        "Product pack size": master_dict["Product pack size"],
        "Corporation": company_name,
        "Molecule": molecule,
        "Shortage Status": shortage_status,
        "Date reported": report_date,
        "Start Date": master_dict["start_date"],
        "End Date": master_dict["end_date"],
        "Date last updated": master_dict["start_date"],
        "Relevant unit numbers": "",
        "Alternatives available": master_dict["Alternatives available"],
        "Shortage impact rating": "",
        "Additional Commentary": master_dict["Additional Commentary"]
    }

    await write_dict_to_xlsm(data)
    print(data)
    print("-" * 50)


async def run():
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        context = await browser.new_context(
            viewport={"width": 1280, "height": 3000}  # Bigger height to load more content
        )

        page = await context.new_page()

        await page.goto("https://cima.aemps.es/cima/publico/listadesabastecimiento.html", timeout=60000)

        # Scroll to load all items
        await auto_scroll(page)

        # Get all list items
        items = await page.query_selector_all("div#resultlist div.list-group-item.row div.titleDesabast")

        print(f"Found {len(items)} items")

        for i in range(len(items)):
            # Refetch items each time because DOM changes after navigation
            items = await page.query_selector_all("div#resultlist div.list-group-item.row div.titleDesabast")
            item = items[i]

            main_title = (await item.inner_text()).strip() if item else ""

            side_box = await page.query_selector_all("div#resultlist div.list-group-item.row")
            side_box = side_box[i]
            master_dict = {}
            master_dict["start_date"], master_dict["end_date"] = await extract_dates(side_box)

            # Pack size
            master_dict["Product pack size"] = await get_text_after_last_comma(main_title)

            # Alternatives
            alternatives_tag = await page.query_selector("div.list-group-item-text-normal")
            alternate = (await alternatives_tag.inner_text()).strip() if alternatives_tag else ""
            master_dict["Alternatives available"] = "No"
            if "Existe/n otro/s medicamento/s" in alternate:
                master_dict["Alternatives available"] = "Yes"

            master_dict["Additional Commentary"] = alternate

            await item.click()
            await page.wait_for_selector("h1#nombreMedicamento", timeout=10000)

            await scrape_detail(page, master_dict)

            # Go back
            await page.go_back()
            await page.wait_for_selector("div#resultlist div.list-group-item.row div.titleDesabast", timeout=10000)

        await browser.close()


asyncio.run(run())
